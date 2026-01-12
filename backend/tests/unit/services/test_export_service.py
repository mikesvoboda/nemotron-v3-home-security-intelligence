"""Unit tests for the export service (NEM-2088).

Tests verify CSV and Excel export functionality including:
1. Format detection from Accept headers
2. CSV generation with proper escaping
3. Excel generation with formatting
4. CSV injection protection
5. Filename generation
"""

from datetime import UTC, datetime

import pytest

from backend.services.export_service import (
    ACCEPT_HEADER_MAPPING,
    EXPORT_COLUMNS,
    EXPORT_EXTENSIONS,
    EXPORT_MIME_TYPES,
    EXTENDED_EXPORT_COLUMNS,
    EventExportRow,
    ExportFormat,
    ExportService,
    events_to_csv,
    events_to_excel,
    format_export_value,
    generate_export_filename,
    get_export_service,
    parse_accept_header,
    reset_export_service,
    sanitize_export_value,
)


class TestSanitizeExportValue:
    """Tests for CSV injection protection in export values."""

    def test_equals_sign_sanitized(self):
        """Test that values starting with = are sanitized."""
        malicious = '=HYPERLINK("http://evil.com","Click")'
        sanitized = sanitize_export_value(malicious)

        assert not sanitized.startswith("=")
        assert sanitized.startswith("'=")

    def test_plus_sign_sanitized(self):
        """Test that values starting with + are sanitized."""
        malicious = "+1+1"
        sanitized = sanitize_export_value(malicious)

        assert not sanitized.startswith("+")
        assert sanitized.startswith("'+")

    def test_minus_sign_sanitized(self):
        """Test that values starting with - are sanitized."""
        malicious = "-1+1"
        sanitized = sanitize_export_value(malicious)

        assert not sanitized.startswith("-")
        assert sanitized.startswith("'-")

    def test_at_sign_sanitized(self):
        """Test that values starting with @ are sanitized."""
        malicious = "@SUM(A1:A10)"
        sanitized = sanitize_export_value(malicious)

        assert not sanitized.startswith("@")
        assert sanitized.startswith("'@")

    def test_tab_sanitized(self):
        """Test that values starting with tab are sanitized."""
        malicious = "\t=cmd|'/C calc'!A0"
        sanitized = sanitize_export_value(malicious)

        assert not sanitized.startswith("\t")
        assert sanitized.startswith("'\t")

    def test_carriage_return_sanitized(self):
        """Test that values starting with carriage return are sanitized."""
        malicious = "\r=cmd|'/C calc'!A0"
        sanitized = sanitize_export_value(malicious)

        assert not sanitized.startswith("\r")
        assert sanitized.startswith("'\r")

    def test_normal_value_unchanged(self):
        """Test that normal values are not modified."""
        normal = "Person detected near entrance"
        sanitized = sanitize_export_value(normal)

        assert sanitized == normal

    def test_empty_string_unchanged(self):
        """Test that empty strings are handled correctly."""
        assert sanitize_export_value("") == ""

    def test_none_returns_empty_string(self):
        """Test that None returns empty string."""
        assert sanitize_export_value(None) == ""

    def test_value_with_special_char_in_middle_unchanged(self):
        """Test that values with special chars in middle are not modified."""
        value = "Person detected - 10:00 AM"
        sanitized = sanitize_export_value(value)

        assert sanitized == value


class TestParseAcceptHeader:
    """Tests for Accept header parsing."""

    def test_csv_text_csv(self):
        """Test text/csv maps to CSV format."""
        assert parse_accept_header("text/csv") == ExportFormat.CSV

    def test_csv_application_csv(self):
        """Test application/csv maps to CSV format."""
        assert parse_accept_header("application/csv") == ExportFormat.CSV

    def test_excel_openxml(self):
        """Test XLSX MIME type maps to Excel format."""
        accept = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        assert parse_accept_header(accept) == ExportFormat.EXCEL

    def test_excel_ms_excel(self):
        """Test application/vnd.ms-excel maps to Excel format."""
        assert parse_accept_header("application/vnd.ms-excel") == ExportFormat.EXCEL

    def test_excel_xlsx(self):
        """Test application/xlsx maps to Excel format."""
        assert parse_accept_header("application/xlsx") == ExportFormat.EXCEL

    def test_none_defaults_to_csv(self):
        """Test that None Accept header defaults to CSV."""
        assert parse_accept_header(None) == ExportFormat.CSV

    def test_empty_defaults_to_csv(self):
        """Test that empty Accept header defaults to CSV."""
        assert parse_accept_header("") == ExportFormat.CSV

    def test_unknown_defaults_to_csv(self):
        """Test that unknown MIME type defaults to CSV."""
        assert parse_accept_header("text/html") == ExportFormat.CSV
        assert parse_accept_header("application/xml") == ExportFormat.CSV
        # application/json also defaults to CSV since JSON export isn't supported
        assert parse_accept_header("application/json") == ExportFormat.CSV

    def test_wildcard_defaults_to_csv(self):
        """Test that */* defaults to CSV."""
        assert parse_accept_header("*/*") == ExportFormat.CSV

    def test_text_wildcard_defaults_to_csv(self):
        """Test that text/* defaults to CSV."""
        assert parse_accept_header("text/*") == ExportFormat.CSV

    def test_accept_header_with_quality_values(self):
        """Test parsing Accept header with quality values."""
        accept = "text/csv;q=0.9, application/json;q=0.8"
        assert parse_accept_header(accept) == ExportFormat.CSV

    def test_accept_header_multiple_types_prefers_first_match(self):
        """Test that first matching type is used."""
        accept = "text/csv, application/xlsx"
        assert parse_accept_header(accept) == ExportFormat.CSV

    def test_accept_header_excel_first(self):
        """Test Excel format when listed first."""
        accept = "application/xlsx, text/csv"
        assert parse_accept_header(accept) == ExportFormat.EXCEL

    def test_case_insensitive(self):
        """Test that Accept header parsing is case-insensitive."""
        assert parse_accept_header("TEXT/CSV") == ExportFormat.CSV
        assert parse_accept_header("Application/XLSX") == ExportFormat.EXCEL


class TestGenerateExportFilename:
    """Tests for filename generation."""

    def test_csv_extension(self):
        """Test CSV files get .csv extension."""
        filename = generate_export_filename("events_export", ExportFormat.CSV)

        assert filename.startswith("events_export_")
        assert filename.endswith(".csv")

    def test_excel_extension(self):
        """Test Excel files get .xlsx extension."""
        filename = generate_export_filename("events_export", ExportFormat.EXCEL)

        assert filename.startswith("events_export_")
        assert filename.endswith(".xlsx")

    def test_timestamp_format(self):
        """Test filename includes timestamp in expected format."""
        filename = generate_export_filename("test", ExportFormat.CSV)

        # Extract timestamp part: test_YYYYMMDD_HHMMSS.csv
        parts = filename.replace(".csv", "").split("_")
        assert len(parts) >= 3

        # Verify timestamp format
        date_part = parts[1]
        time_part = parts[2]
        assert len(date_part) == 8  # YYYYMMDD
        assert len(time_part) == 6  # HHMMSS


class TestEventExportRow:
    """Tests for EventExportRow dataclass."""

    def test_create_minimal_row(self):
        """Test creating row with required fields only."""
        row = EventExportRow(
            event_id=1,
            camera_name="Front Door",
            started_at=datetime(2024, 1, 15, 10, 30, 0, tzinfo=UTC),
            ended_at=None,
            risk_score=75,
            risk_level="high",
            summary="Person detected",
            detection_count=3,
            reviewed=False,
        )

        assert row.event_id == 1
        assert row.camera_name == "Front Door"
        assert row.detection_count == 3

    def test_create_full_row(self):
        """Test creating row with all fields."""
        row = EventExportRow(
            event_id=1,
            camera_name="Front Door",
            started_at=datetime(2024, 1, 15, 10, 30, 0, tzinfo=UTC),
            ended_at=datetime(2024, 1, 15, 10, 31, 30, tzinfo=UTC),
            risk_score=75,
            risk_level="high",
            summary="Person detected",
            detection_count=3,
            reviewed=True,
            object_types="person,vehicle",
            reasoning="Multiple persons detected near entrance",
        )

        assert row.object_types == "person,vehicle"
        assert row.reasoning == "Multiple persons detected near entrance"


class TestFormatExportValue:
    """Tests for export value formatting."""

    def test_format_datetime(self):
        """Test datetime formatting as ISO string."""
        row = EventExportRow(
            event_id=1,
            camera_name="Test",
            started_at=datetime(2024, 1, 15, 10, 30, 0, tzinfo=UTC),
            ended_at=None,
            risk_score=50,
            risk_level="medium",
            summary="Test",
            detection_count=1,
            reviewed=False,
        )

        value = format_export_value(row, "started_at")
        assert "2024-01-15" in value
        assert "10:30:00" in value

    def test_format_boolean_true(self):
        """Test boolean True formats as Yes."""
        row = EventExportRow(
            event_id=1,
            camera_name="Test",
            started_at=datetime(2024, 1, 15, 10, 30, 0, tzinfo=UTC),
            ended_at=None,
            risk_score=50,
            risk_level="medium",
            summary="Test",
            detection_count=1,
            reviewed=True,
        )

        value = format_export_value(row, "reviewed")
        assert value == "Yes"

    def test_format_boolean_false(self):
        """Test boolean False formats as No."""
        row = EventExportRow(
            event_id=1,
            camera_name="Test",
            started_at=datetime(2024, 1, 15, 10, 30, 0, tzinfo=UTC),
            ended_at=None,
            risk_score=50,
            risk_level="medium",
            summary="Test",
            detection_count=1,
            reviewed=False,
        )

        value = format_export_value(row, "reviewed")
        assert value == "No"

    def test_format_integer(self):
        """Test integer formatting."""
        row = EventExportRow(
            event_id=42,
            camera_name="Test",
            started_at=datetime(2024, 1, 15, 10, 30, 0, tzinfo=UTC),
            ended_at=None,
            risk_score=75,
            risk_level="high",
            summary="Test",
            detection_count=5,
            reviewed=False,
        )

        assert format_export_value(row, "event_id") == "42"
        assert format_export_value(row, "risk_score") == "75"

    def test_format_none(self):
        """Test None formatting as empty string."""
        row = EventExportRow(
            event_id=1,
            camera_name="Test",
            started_at=datetime(2024, 1, 15, 10, 30, 0, tzinfo=UTC),
            ended_at=None,
            risk_score=None,
            risk_level=None,
            summary=None,
            detection_count=1,
            reviewed=False,
        )

        assert format_export_value(row, "ended_at") == ""
        assert format_export_value(row, "risk_score") == ""
        assert format_export_value(row, "summary") == ""

    def test_format_string_with_injection_char(self):
        """Test string with injection character is sanitized."""
        row = EventExportRow(
            event_id=1,
            camera_name="Test",
            started_at=datetime(2024, 1, 15, 10, 30, 0, tzinfo=UTC),
            ended_at=None,
            risk_score=50,
            risk_level="medium",
            summary="=HYPERLINK(...)",
            detection_count=1,
            reviewed=False,
        )

        value = format_export_value(row, "summary")
        assert value.startswith("'=")


class TestEventsToCSV:
    """Tests for CSV generation."""

    @pytest.fixture
    def sample_events(self) -> list[EventExportRow]:
        """Create sample events for testing."""
        return [
            EventExportRow(
                event_id=1,
                camera_name="Front Door",
                started_at=datetime(2024, 1, 15, 10, 30, 0, tzinfo=UTC),
                ended_at=datetime(2024, 1, 15, 10, 31, 30, tzinfo=UTC),
                risk_score=75,
                risk_level="high",
                summary="Person detected at entrance",
                detection_count=3,
                reviewed=False,
            ),
            EventExportRow(
                event_id=2,
                camera_name="Back Yard",
                started_at=datetime(2024, 1, 15, 11, 0, 0, tzinfo=UTC),
                ended_at=None,
                risk_score=25,
                risk_level="low",
                summary="Cat in garden",
                detection_count=1,
                reviewed=True,
            ),
        ]

    def test_csv_has_header_row(self, sample_events: list[EventExportRow]):
        """Test CSV output includes header row."""
        csv_content = events_to_csv(sample_events)
        lines = csv_content.strip().split("\n")

        # First line should be header
        header = lines[0]
        assert "Event ID" in header
        assert "Camera" in header
        assert "Started At" in header
        assert "Risk Score" in header

    def test_csv_has_data_rows(self, sample_events: list[EventExportRow]):
        """Test CSV output includes data rows."""
        csv_content = events_to_csv(sample_events)
        lines = csv_content.strip().split("\n")

        # Should have header + 2 data rows
        assert len(lines) == 3

    def test_csv_contains_event_data(self, sample_events: list[EventExportRow]):
        """Test CSV contains expected event data."""
        csv_content = events_to_csv(sample_events)

        assert "Front Door" in csv_content
        assert "Back Yard" in csv_content
        assert "Person detected at entrance" in csv_content
        assert "Cat in garden" in csv_content

    def test_csv_empty_list(self):
        """Test CSV with empty event list."""
        csv_content = events_to_csv([])
        lines = csv_content.strip().split("\n")

        # Should have header only
        assert len(lines) == 1
        assert "Event ID" in lines[0]

    def test_csv_custom_columns(self, sample_events: list[EventExportRow]):
        """Test CSV with custom column selection."""
        columns = [
            ("event_id", "ID"),
            ("camera_name", "Camera"),
            ("risk_score", "Score"),
        ]

        csv_content = events_to_csv(sample_events, columns=columns)
        lines = csv_content.strip().split("\n")

        # Header should only have custom columns
        header = lines[0]
        assert "ID" in header
        assert "Camera" in header
        assert "Score" in header
        assert "Summary" not in header


class TestEventsToExcel:
    """Tests for Excel generation."""

    @pytest.fixture
    def sample_events(self) -> list[EventExportRow]:
        """Create sample events for testing."""
        return [
            EventExportRow(
                event_id=1,
                camera_name="Front Door",
                started_at=datetime(2024, 1, 15, 10, 30, 0, tzinfo=UTC),
                ended_at=datetime(2024, 1, 15, 10, 31, 30, tzinfo=UTC),
                risk_score=75,
                risk_level="high",
                summary="Person detected at entrance",
                detection_count=3,
                reviewed=False,
            ),
            EventExportRow(
                event_id=2,
                camera_name="Back Yard",
                started_at=datetime(2024, 1, 15, 11, 0, 0, tzinfo=UTC),
                ended_at=None,
                risk_score=25,
                risk_level="low",
                summary="Cat in garden",
                detection_count=1,
                reviewed=True,
            ),
        ]

    def test_excel_returns_bytes(self, sample_events: list[EventExportRow]):
        """Test Excel export returns bytes."""
        content = events_to_excel(sample_events)

        assert isinstance(content, bytes)
        assert len(content) > 0

    def test_excel_has_xlsx_signature(self, sample_events: list[EventExportRow]):
        """Test Excel file has XLSX signature (ZIP format)."""
        content = events_to_excel(sample_events)

        # XLSX files are ZIP archives, start with PK signature
        assert content[:2] == b"PK"

    def test_excel_can_be_read_by_openpyxl(self, sample_events: list[EventExportRow]):
        """Test Excel file can be read back by openpyxl."""
        import io

        from openpyxl import load_workbook

        content = events_to_excel(sample_events)
        wb = load_workbook(io.BytesIO(content))

        assert wb.active is not None
        ws = wb.active
        assert ws.title == "Events"

    def test_excel_contains_header_row(self, sample_events: list[EventExportRow]):
        """Test Excel file contains header row."""
        import io

        from openpyxl import load_workbook

        content = events_to_excel(sample_events)
        wb = load_workbook(io.BytesIO(content))
        ws = wb.active

        # Check header row (row 1)
        assert ws.cell(row=1, column=1).value == "Event ID"
        assert ws.cell(row=1, column=2).value == "Camera"

    def test_excel_contains_data(self, sample_events: list[EventExportRow]):
        """Test Excel file contains event data."""
        import io

        from openpyxl import load_workbook

        content = events_to_excel(sample_events)
        wb = load_workbook(io.BytesIO(content))
        ws = wb.active

        # Check first data row (row 2)
        assert ws.cell(row=2, column=1).value == 1  # event_id
        assert ws.cell(row=2, column=2).value == "Front Door"  # camera_name

        # Check second data row (row 3)
        assert ws.cell(row=3, column=1).value == 2
        assert ws.cell(row=3, column=2).value == "Back Yard"

    def test_excel_custom_sheet_name(self, sample_events: list[EventExportRow]):
        """Test Excel file with custom sheet name."""
        import io

        from openpyxl import load_workbook

        content = events_to_excel(sample_events, sheet_name="Security Events")
        wb = load_workbook(io.BytesIO(content))
        ws = wb.active

        assert ws.title == "Security Events"

    def test_excel_empty_list(self):
        """Test Excel with empty event list."""
        import io

        from openpyxl import load_workbook

        content = events_to_excel([])
        wb = load_workbook(io.BytesIO(content))
        ws = wb.active

        # Should have header only
        assert ws.cell(row=1, column=1).value == "Event ID"
        assert ws.cell(row=2, column=1).value is None

    def test_excel_handles_none_values(self):
        """Test Excel handles None values correctly."""
        import io

        from openpyxl import load_workbook

        events = [
            EventExportRow(
                event_id=1,
                camera_name="Test",
                started_at=datetime(2024, 1, 15, 10, 30, 0, tzinfo=UTC),
                ended_at=None,
                risk_score=None,
                risk_level=None,
                summary=None,
                detection_count=0,
                reviewed=False,
            ),
        ]

        content = events_to_excel(events)
        wb = load_workbook(io.BytesIO(content))
        ws = wb.active

        # None values should be empty strings
        assert ws.cell(row=2, column=4).value in (None, "")  # ended_at
        assert ws.cell(row=2, column=5).value in (None, "")  # risk_score


class TestExportService:
    """Tests for ExportService class."""

    @pytest.fixture(autouse=True)
    def reset_service(self):
        """Reset service singleton before each test."""
        reset_export_service()
        yield
        reset_export_service()

    def test_get_export_service_singleton(self):
        """Test get_export_service returns singleton."""
        service1 = get_export_service()
        service2 = get_export_service()

        assert service1 is service2

    def test_get_export_format(self):
        """Test format detection from Accept header."""
        service = ExportService()

        assert service.get_export_format("text/csv") == ExportFormat.CSV
        assert service.get_export_format("application/xlsx") == ExportFormat.EXCEL

    def test_get_content_type_csv(self):
        """Test content type for CSV format."""
        service = ExportService()

        content_type = service.get_content_type(ExportFormat.CSV)
        assert content_type == "text/csv"

    def test_get_content_type_excel(self):
        """Test content type for Excel format."""
        service = ExportService()

        content_type = service.get_content_type(ExportFormat.EXCEL)
        assert content_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    def test_get_filename(self):
        """Test filename generation."""
        service = ExportService()

        csv_filename = service.get_filename("events", ExportFormat.CSV)
        excel_filename = service.get_filename("events", ExportFormat.EXCEL)

        assert csv_filename.endswith(".csv")
        assert excel_filename.endswith(".xlsx")

    def test_export_events_csv(self):
        """Test exporting events as CSV."""
        service = ExportService()
        events = [
            EventExportRow(
                event_id=1,
                camera_name="Test",
                started_at=datetime(2024, 1, 15, 10, 30, 0, tzinfo=UTC),
                ended_at=None,
                risk_score=50,
                risk_level="medium",
                summary="Test event",
                detection_count=1,
                reviewed=False,
            ),
        ]

        content = service.export_events(events, ExportFormat.CSV)

        assert isinstance(content, str)
        assert "Test" in content
        assert "Event ID" in content

    def test_export_events_excel(self):
        """Test exporting events as Excel."""
        service = ExportService()
        events = [
            EventExportRow(
                event_id=1,
                camera_name="Test",
                started_at=datetime(2024, 1, 15, 10, 30, 0, tzinfo=UTC),
                ended_at=None,
                risk_score=50,
                risk_level="medium",
                summary="Test event",
                detection_count=1,
                reviewed=False,
            ),
        ]

        content = service.export_events(events, ExportFormat.EXCEL)

        assert isinstance(content, bytes)
        assert content[:2] == b"PK"  # ZIP signature

    def test_content_disposition_attachment(self):
        """Test Content-Disposition header for attachment."""
        service = ExportService()

        header = service.get_content_disposition_header("events.csv")
        assert header == 'attachment; filename="events.csv"'

    def test_content_disposition_inline(self):
        """Test Content-Disposition header for inline."""
        service = ExportService()

        header = service.get_content_disposition_header("events.csv", inline=True)
        assert header == 'inline; filename="events.csv"'


class TestEventsToCSVStreaming:
    """Tests for CSV streaming generation."""

    @pytest.fixture
    def sample_events(self) -> list[EventExportRow]:
        """Create sample events for testing."""
        return [
            EventExportRow(
                event_id=1,
                camera_name="Front Door",
                started_at=datetime(2024, 1, 15, 10, 30, 0, tzinfo=UTC),
                ended_at=datetime(2024, 1, 15, 10, 31, 30, tzinfo=UTC),
                risk_score=75,
                risk_level="high",
                summary="Person detected at entrance",
                detection_count=3,
                reviewed=False,
            ),
            EventExportRow(
                event_id=2,
                camera_name="Back Yard",
                started_at=datetime(2024, 1, 15, 11, 0, 0, tzinfo=UTC),
                ended_at=None,
                risk_score=25,
                risk_level="low",
                summary="Cat in garden",
                detection_count=1,
                reviewed=True,
            ),
        ]

    def test_streaming_yields_chunks(self, sample_events: list[EventExportRow]):
        """Test that streaming yields chunks of CSV data."""
        from backend.services.export_service import events_to_csv_streaming

        chunks = list(events_to_csv_streaming(sample_events))

        # Should have at least header + 2 data rows
        assert len(chunks) >= 3

    def test_streaming_first_chunk_is_header(self, sample_events: list[EventExportRow]):
        """Test that first chunk contains header row."""
        from backend.services.export_service import events_to_csv_streaming

        chunks = list(events_to_csv_streaming(sample_events))
        first_chunk = chunks[0]

        assert "Event ID" in first_chunk
        assert "Camera" in first_chunk

    def test_streaming_contains_data(self, sample_events: list[EventExportRow]):
        """Test that streaming contains event data."""
        from backend.services.export_service import events_to_csv_streaming

        chunks = list(events_to_csv_streaming(sample_events))
        all_content = "".join(chunks)

        assert "Front Door" in all_content
        assert "Back Yard" in all_content

    def test_streaming_with_empty_list(self):
        """Test streaming with empty event list."""
        from backend.services.export_service import events_to_csv_streaming

        chunks = list(events_to_csv_streaming([]))

        # Should only yield header
        assert len(chunks) == 1
        assert "Event ID" in chunks[0]

    def test_streaming_with_custom_columns(self, sample_events: list[EventExportRow]):
        """Test streaming with custom columns."""
        from backend.services.export_service import events_to_csv_streaming

        columns = [
            ("event_id", "ID"),
            ("camera_name", "Camera"),
        ]

        chunks = list(events_to_csv_streaming(sample_events, columns=columns))
        first_chunk = chunks[0]

        assert "ID" in first_chunk
        assert "Camera" in first_chunk
        assert "Risk Score" not in first_chunk


@pytest.mark.asyncio
class TestExportServiceWithProgress:
    """Tests for export service with progress tracking."""

    @pytest.fixture
    async def mock_db(self):
        """Create a mock database session."""
        from unittest.mock import AsyncMock

        db = AsyncMock()
        db.execute = AsyncMock()
        return db

    @pytest.fixture
    def mock_job_tracker(self):
        """Create a mock job tracker."""
        from unittest.mock import MagicMock

        tracker = MagicMock()
        tracker.update_progress = MagicMock()
        return tracker

    async def test_export_with_progress_no_db_raises_error(self):
        """Test that export with progress raises error without database."""
        from unittest.mock import MagicMock

        service = ExportService()
        tracker = MagicMock()

        with pytest.raises(ValueError, match="Database session required"):
            await service.export_events_with_progress(
                job_id="test-job",
                job_tracker=tracker,
                export_format="csv",
            )

    async def test_export_with_progress_empty_results(self, mock_db, mock_job_tracker):
        """Test export with progress when no events match filters."""
        from unittest.mock import MagicMock

        # Mock empty count result
        count_result = MagicMock()
        count_result.scalar.return_value = 0
        mock_db.execute.return_value = count_result

        service = ExportService(db=mock_db)

        result = await service.export_events_with_progress(
            job_id="test-job",
            job_tracker=mock_job_tracker,
            export_format="csv",
        )

        # Should create empty export
        assert result["event_count"] == 0
        assert result["format"] == "csv"
        assert "file_path" in result

        # Should update progress
        mock_job_tracker.update_progress.assert_called()

    async def test_export_with_progress_invalid_format(self, mock_db, mock_job_tracker):
        """Test export with progress with invalid format."""
        from unittest.mock import MagicMock

        # Mock empty count
        count_result = MagicMock()
        count_result.scalar.return_value = 0
        mock_db.execute.return_value = count_result

        service = ExportService(db=mock_db)

        with pytest.raises(ValueError, match="Unsupported export format"):
            await service.export_events_with_progress(
                job_id="test-job",
                job_tracker=mock_job_tracker,
                export_format="invalid",
            )

    async def test_export_with_progress_csv_format(self, mock_db, mock_job_tracker):
        """Test export with progress for CSV format."""
        from unittest.mock import AsyncMock, MagicMock

        # Mock event query result using MagicMock
        event = MagicMock()
        event.id = 1
        event.camera_id = "cam-1"
        event.started_at = datetime(2024, 1, 15, 10, 30, 0, tzinfo=UTC)
        event.ended_at = datetime(2024, 1, 15, 10, 31, 30, tzinfo=UTC)
        event.risk_score = 75
        event.risk_level = "high"
        event.summary = "Test event"
        event.detection_count = 3
        event.reviewed = False
        event.object_types = None
        event.reasoning = None

        event_result = MagicMock()
        event_result.scalars.return_value.all.return_value = [event]

        # Mock camera query result
        camera_result = MagicMock()
        camera_result.scalar.return_value = "Front Door"

        # Mock count result
        count_result = MagicMock()
        count_result.scalar.return_value = 1

        # Setup mock side_effect to return different results per call
        # Order: count query, events query, camera query (for each event)
        mock_db.execute = AsyncMock(side_effect=[count_result, event_result, camera_result])

        service = ExportService(db=mock_db)

        result = await service.export_events_with_progress(
            job_id="test-job",
            job_tracker=mock_job_tracker,
            export_format="csv",
        )

        assert result["format"] == "csv"
        assert result["event_count"] == 1
        assert result["file_path"].endswith(".csv")
        assert result["file_size"] > 0

    async def test_export_with_progress_json_format(self, mock_db, mock_job_tracker):
        """Test export with progress for JSON format."""
        from unittest.mock import AsyncMock, MagicMock

        # Mock event query result
        event = MagicMock()
        event.id = 1
        event.camera_id = "cam-1"
        event.started_at = datetime(2024, 1, 15, 10, 30, 0, tzinfo=UTC)
        event.ended_at = None
        event.risk_score = 75
        event.risk_level = "high"
        event.summary = "Test"
        event.detection_count = 1
        event.reviewed = False
        event.object_types = None
        event.reasoning = None

        event_result = MagicMock()
        event_result.scalars.return_value.all.return_value = [event]

        # Mock camera query result
        camera_result = MagicMock()
        camera_result.scalar.return_value = "Camera"

        # Mock count result
        count_result = MagicMock()
        count_result.scalar.return_value = 1

        mock_db.execute = AsyncMock(side_effect=[count_result, event_result, camera_result])

        service = ExportService(db=mock_db)

        result = await service.export_events_with_progress(
            job_id="test-job",
            job_tracker=mock_job_tracker,
            export_format="json",
        )

        assert result["format"] == "json"
        assert result["file_path"].endswith(".json")

    async def test_export_with_progress_zip_format(self, mock_db, mock_job_tracker):
        """Test export with progress for ZIP format."""
        from unittest.mock import AsyncMock, MagicMock

        # Mock event query result
        event = MagicMock()
        event.id = 1
        event.camera_id = "cam-1"
        event.started_at = datetime(2024, 1, 15, 10, 30, 0, tzinfo=UTC)
        event.ended_at = None
        event.risk_score = 50
        event.risk_level = "medium"
        event.summary = "Test"
        event.detection_count = 1
        event.reviewed = False
        event.object_types = None
        event.reasoning = None

        event_result = MagicMock()
        event_result.scalars.return_value.all.return_value = [event]

        # Mock camera query result
        camera_result = MagicMock()
        camera_result.scalar.return_value = "Camera"

        # Mock count result
        count_result = MagicMock()
        count_result.scalar.return_value = 1

        mock_db.execute = AsyncMock(side_effect=[count_result, event_result, camera_result])

        service = ExportService(db=mock_db)

        result = await service.export_events_with_progress(
            job_id="test-job",
            job_tracker=mock_job_tracker,
            export_format="zip",
        )

        assert result["format"] == "zip"
        assert result["file_path"].endswith(".zip")

    async def test_export_with_progress_applies_camera_filter(self, mock_db, mock_job_tracker):
        """Test that camera filter is applied in query."""
        from unittest.mock import MagicMock

        # Mock empty count
        count_result = MagicMock()
        count_result.scalar.return_value = 0
        mock_db.execute.return_value = count_result

        service = ExportService(db=mock_db)

        await service.export_events_with_progress(
            job_id="test-job",
            job_tracker=mock_job_tracker,
            export_format="csv",
            camera_id="cam-123",
        )

        # Should have called execute (filter logic is in SQL)
        assert mock_db.execute.called

    async def test_export_with_progress_applies_risk_level_filter(self, mock_db, mock_job_tracker):
        """Test that risk level filter is applied in query."""
        from unittest.mock import MagicMock

        # Mock empty count
        count_result = MagicMock()
        count_result.scalar.return_value = 0
        mock_db.execute.return_value = count_result

        service = ExportService(db=mock_db)

        await service.export_events_with_progress(
            job_id="test-job",
            job_tracker=mock_job_tracker,
            export_format="csv",
            risk_level="high",
        )

        assert mock_db.execute.called

    async def test_export_with_progress_applies_date_filters(self, mock_db, mock_job_tracker):
        """Test that date filters are applied in query."""
        from unittest.mock import MagicMock

        # Mock empty count
        count_result = MagicMock()
        count_result.scalar.return_value = 0
        mock_db.execute.return_value = count_result

        service = ExportService(db=mock_db)

        await service.export_events_with_progress(
            job_id="test-job",
            job_tracker=mock_job_tracker,
            export_format="csv",
            start_date="2024-01-15T00:00:00Z",
            end_date="2024-01-16T00:00:00Z",
        )

        assert mock_db.execute.called

    async def test_export_with_progress_applies_reviewed_filter(self, mock_db, mock_job_tracker):
        """Test that reviewed filter is applied in query."""
        from unittest.mock import MagicMock

        # Mock empty count
        count_result = MagicMock()
        count_result.scalar.return_value = 0
        mock_db.execute.return_value = count_result

        service = ExportService(db=mock_db)

        await service.export_events_with_progress(
            job_id="test-job",
            job_tracker=mock_job_tracker,
            export_format="csv",
            reviewed=True,
        )

        assert mock_db.execute.called


@pytest.mark.asyncio
class TestCreateEmptyExport:
    """Tests for _create_empty_export method."""

    async def test_create_empty_csv(self):
        """Test creating empty CSV export."""
        service = ExportService()

        result = await service._create_empty_export("csv")

        assert result["format"] == "csv"
        assert result["event_count"] == 0
        assert result["file_path"].endswith(".csv")
        assert result["file_size"] > 0

    async def test_create_empty_json(self):
        """Test creating empty JSON export."""
        service = ExportService()

        result = await service._create_empty_export("json")

        assert result["format"] == "json"
        assert result["event_count"] == 0
        assert result["file_path"].endswith(".json")
        assert result["file_size"] > 0

    async def test_create_empty_zip(self):
        """Test creating empty ZIP export."""
        service = ExportService()

        result = await service._create_empty_export("zip")

        assert result["format"] == "zip"
        assert result["event_count"] == 0
        assert result["file_path"].endswith(".zip")
        assert result["file_size"] > 0

    async def test_create_empty_invalid_format(self):
        """Test creating empty export with invalid format."""
        service = ExportService()

        with pytest.raises(ValueError, match="Unsupported export format"):
            await service._create_empty_export("invalid")


@pytest.mark.asyncio
class TestExportServiceWithWebSocket:
    """Tests for export service with WebSocket progress reporting."""

    @pytest.fixture
    async def mock_db(self):
        """Create a mock database session."""
        from unittest.mock import AsyncMock

        db = AsyncMock()
        db.execute = AsyncMock()
        return db

    @pytest.fixture
    def mock_progress_reporter(self):
        """Create a mock progress reporter."""
        from unittest.mock import AsyncMock, MagicMock

        reporter = MagicMock()
        reporter.start = AsyncMock()
        reporter.report_progress = AsyncMock()
        reporter.complete = AsyncMock()
        reporter.fail = AsyncMock()
        reporter.job_id = "test-job-123"
        reporter.duration_seconds = 1.23
        return reporter

    async def test_export_with_websocket_no_db_raises_error(self, mock_progress_reporter):
        """Test that export with websocket raises error without database."""
        service = ExportService()

        with pytest.raises(ValueError, match="Database session required"):
            await service.export_events_with_websocket(
                progress_reporter=mock_progress_reporter,
                export_format="csv",
            )

    async def test_export_with_websocket_starts_reporter(self, mock_db, mock_progress_reporter):
        """Test that reporter is started with metadata."""
        from unittest.mock import MagicMock

        # Mock empty count
        count_result = MagicMock()
        count_result.scalar.return_value = 0
        mock_db.execute.return_value = count_result

        service = ExportService(db=mock_db)

        await service.export_events_with_websocket(
            progress_reporter=mock_progress_reporter,
            export_format="csv",
        )

        # Verify reporter was started with metadata
        mock_progress_reporter.start.assert_called_once()
        call_args = mock_progress_reporter.start.call_args
        assert "metadata" in call_args[1]
        assert call_args[1]["metadata"]["export_format"] == "csv"

    async def test_export_with_websocket_completes_successfully(
        self, mock_db, mock_progress_reporter
    ):
        """Test that successful export completes the reporter."""
        from unittest.mock import MagicMock

        # Mock empty count
        count_result = MagicMock()
        count_result.scalar.return_value = 0
        mock_db.execute.return_value = count_result

        service = ExportService(db=mock_db)

        result = await service.export_events_with_websocket(
            progress_reporter=mock_progress_reporter,
            export_format="csv",
        )

        # Verify reporter was completed
        mock_progress_reporter.complete.assert_called_once()
        assert result["event_count"] == 0

    async def test_export_with_websocket_reports_progress(self, mock_db, mock_progress_reporter):
        """Test that progress is reported during export."""
        from unittest.mock import AsyncMock, MagicMock

        # Mock event
        event = MagicMock()
        event.id = 1
        event.camera_id = "cam-1"
        event.started_at = datetime(2024, 1, 15, 10, 30, 0, tzinfo=UTC)
        event.ended_at = None
        event.risk_score = 75
        event.risk_level = "high"
        event.summary = "Test"
        event.detection_count = 1
        event.reviewed = False
        event.object_types = None
        event.reasoning = None

        event_result = MagicMock()
        event_result.scalars.return_value.all.return_value = [event]

        camera_result = MagicMock()
        camera_result.scalar.return_value = "Camera"

        # Mock count result
        count_result = MagicMock()
        count_result.scalar.return_value = 1

        mock_db.execute = AsyncMock(side_effect=[count_result, event_result, camera_result])

        service = ExportService(db=mock_db)

        await service.export_events_with_websocket(
            progress_reporter=mock_progress_reporter,
            export_format="csv",
        )

        # Verify progress was reported
        assert mock_progress_reporter.report_progress.call_count > 0

    async def test_export_with_websocket_handles_exception(self, mock_db, mock_progress_reporter):
        """Test that exceptions are handled and reporter fails."""

        # Mock database to raise exception
        mock_db.execute.side_effect = RuntimeError("Database error")

        service = ExportService(db=mock_db)

        with pytest.raises(RuntimeError, match="Database error"):
            await service.export_events_with_websocket(
                progress_reporter=mock_progress_reporter,
                export_format="csv",
            )

        # Verify reporter was failed
        mock_progress_reporter.fail.assert_called_once()
        call_args = mock_progress_reporter.fail.call_args
        assert isinstance(call_args[0][0], RuntimeError)

    async def test_export_with_websocket_invalid_format(self, mock_db, mock_progress_reporter):
        """Test export with websocket with invalid format."""
        from unittest.mock import AsyncMock, MagicMock

        # Mock event
        event = MagicMock()
        event.id = 1
        event.camera_id = "cam-1"
        event.started_at = datetime(2024, 1, 15, 10, 30, 0, tzinfo=UTC)
        event.ended_at = None
        event.risk_score = 75
        event.risk_level = "high"
        event.summary = "Test"
        event.detection_count = 1
        event.reviewed = False
        event.object_types = None
        event.reasoning = None

        event_result = MagicMock()
        event_result.scalars.return_value.all.return_value = [event]

        camera_result = MagicMock()
        camera_result.scalar.return_value = "Camera"

        # Mock count result
        count_result = MagicMock()
        count_result.scalar.return_value = 1

        mock_db.execute = AsyncMock(side_effect=[count_result, event_result, camera_result])

        service = ExportService(db=mock_db)

        with pytest.raises(ValueError, match="Unsupported export format"):
            await service.export_events_with_websocket(
                progress_reporter=mock_progress_reporter,
                export_format="invalid",
            )

        # Verify reporter was failed
        mock_progress_reporter.fail.assert_called_once()

    async def test_export_with_websocket_includes_duration(self, mock_db, mock_progress_reporter):
        """Test that result includes duration from reporter when events exist."""
        from unittest.mock import AsyncMock, MagicMock

        # Mock event
        event = MagicMock()
        event.id = 1
        event.camera_id = "cam-1"
        event.started_at = datetime(2024, 1, 15, 10, 30, 0, tzinfo=UTC)
        event.ended_at = None
        event.risk_score = 75
        event.risk_level = "high"
        event.summary = "Test"
        event.detection_count = 1
        event.reviewed = False
        event.object_types = None
        event.reasoning = None

        event_result = MagicMock()
        event_result.scalars.return_value.all.return_value = [event]

        camera_result = MagicMock()
        camera_result.scalar.return_value = "Camera"

        # Mock count result with 1 event
        count_result = MagicMock()
        count_result.scalar.return_value = 1

        mock_db.execute = AsyncMock(side_effect=[count_result, event_result, camera_result])

        service = ExportService(db=mock_db)

        result = await service.export_events_with_websocket(
            progress_reporter=mock_progress_reporter,
            export_format="csv",
        )

        # Verify duration is included when events exist
        assert "duration_seconds" in result
        assert result["duration_seconds"] == 1.23


class TestExportConstants:
    """Tests for export constant definitions."""

    def test_export_mime_types_defined(self):
        """Test MIME types are defined for all formats."""
        assert ExportFormat.CSV in EXPORT_MIME_TYPES
        assert ExportFormat.EXCEL in EXPORT_MIME_TYPES

    def test_export_extensions_defined(self):
        """Test extensions are defined for all formats."""
        assert ExportFormat.CSV in EXPORT_EXTENSIONS
        assert ExportFormat.EXCEL in EXPORT_EXTENSIONS
        assert EXPORT_EXTENSIONS[ExportFormat.CSV] == ".csv"
        assert EXPORT_EXTENSIONS[ExportFormat.EXCEL] == ".xlsx"

    def test_accept_header_mapping_complete(self):
        """Test Accept header mapping covers common MIME types."""
        assert "text/csv" in ACCEPT_HEADER_MAPPING
        assert "application/csv" in ACCEPT_HEADER_MAPPING
        assert "application/xlsx" in ACCEPT_HEADER_MAPPING
        assert "application/vnd.ms-excel" in ACCEPT_HEADER_MAPPING

    def test_export_columns_defined(self):
        """Test export columns are defined."""
        assert len(EXPORT_COLUMNS) > 0

        # Check required columns exist
        field_names = [col[0] for col in EXPORT_COLUMNS]
        assert "event_id" in field_names
        assert "camera_name" in field_names
        assert "risk_score" in field_names
        assert "reviewed" in field_names

    def test_extended_columns_include_base(self):
        """Test extended columns include base columns."""
        base_fields = {col[0] for col in EXPORT_COLUMNS}
        extended_fields = {col[0] for col in EXTENDED_EXPORT_COLUMNS}

        # All base fields should be in extended
        assert base_fields.issubset(extended_fields)

        # Extended should have additional fields
        assert "object_types" in extended_fields
        assert "reasoning" in extended_fields
