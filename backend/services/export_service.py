"""Export service for generating CSV and Excel exports of event data.

This service provides reusable export functionality for security events,
supporting multiple output formats via content negotiation.

Supported formats:
- CSV (text/csv) - Comma-separated values
- Excel (application/vnd.openxmlformats-officedocument.spreadsheetml.sheet) - XLSX format
- JSON (application/json) - JSON array
- ZIP (application/zip) - Compressed archive with JSON data

Security:
- All string fields are sanitized to prevent CSV injection attacks
- File downloads include Content-Disposition headers with safe filenames

Progress Tracking (NEM-1989):
- Supports background job progress updates via JobTracker
- Updates progress every 100 events for large exports

WebSocket Events (NEM-2380):
- Supports WebSocket event emission via JobProgressReporter
- Emits job.started, job.progress, job.completed, and job.failed events
"""

from __future__ import annotations

import csv
import io
import json
import zipfile
from collections.abc import Iterator, Sequence
from dataclasses import dataclass
from datetime import UTC, datetime
from enum import Enum
from pathlib import Path
from typing import TYPE_CHECKING, Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from backend.core.logging import get_logger

if TYPE_CHECKING:
    from backend.services.job_progress_reporter import JobProgressReporter
    from backend.services.job_tracker import JobTracker

logger = get_logger(__name__)

# Export directory for generated files
EXPORT_DIR = Path("/tmp/exports")  # noqa: S108
PROGRESS_UPDATE_INTERVAL = 100  # Update progress every 100 events


class ExportFormat(str, Enum):
    """Supported export formats."""

    CSV = "csv"
    EXCEL = "excel"


# MIME type mappings for export formats
EXPORT_MIME_TYPES: dict[ExportFormat, str] = {
    ExportFormat.CSV: "text/csv",
    ExportFormat.EXCEL: "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
}

# File extensions for export formats
EXPORT_EXTENSIONS: dict[ExportFormat, str] = {
    ExportFormat.CSV: ".csv",
    ExportFormat.EXCEL: ".xlsx",
}

# Accept header values to format mapping
ACCEPT_HEADER_MAPPING: dict[str, ExportFormat] = {
    "text/csv": ExportFormat.CSV,
    "application/csv": ExportFormat.CSV,
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": ExportFormat.EXCEL,
    "application/vnd.ms-excel": ExportFormat.EXCEL,
    "application/xlsx": ExportFormat.EXCEL,
}

# Characters that trigger formula injection in spreadsheet applications
CSV_INJECTION_PREFIXES = ("=", "+", "-", "@", "\t", "\r")


def sanitize_export_value(value: str | None) -> str:
    """Sanitize a value for safe export to prevent formula injection.

    CSV injection (also known as formula injection) occurs when data
    exported to CSV/Excel is opened in spreadsheet applications. Cells
    starting with certain characters (=, +, -, @, tab, carriage return)
    can be interpreted as formulas, potentially executing malicious code.

    This function prefixes dangerous values with a single quote (')
    which tells spreadsheet applications to treat the cell as text.

    Reference:
    - OWASP CSV Injection: https://owasp.org/www-community/attacks/CSV_Injection

    Args:
        value: The string value to sanitize, or None

    Returns:
        The sanitized string value. Returns empty string if value is None.
    """
    if value is None:
        return ""

    if not value:
        return value

    # Check if the first character is a dangerous injection prefix
    if value[0] in CSV_INJECTION_PREFIXES:
        return f"'{value}"

    return value


@dataclass
class EventExportRow:
    """Represents a single row in an event export.

    This is the canonical structure for export data, used by both
    CSV and Excel export functions.
    """

    event_id: int
    camera_name: str
    started_at: datetime | None
    ended_at: datetime | None
    risk_score: int | None
    risk_level: str | None
    summary: str | None
    detection_count: int
    reviewed: bool
    object_types: str | None = None
    reasoning: str | None = None


# Column definitions for exports
EXPORT_COLUMNS: list[tuple[str, str]] = [
    ("event_id", "Event ID"),
    ("camera_name", "Camera"),
    ("started_at", "Started At"),
    ("ended_at", "Ended At"),
    ("risk_score", "Risk Score"),
    ("risk_level", "Risk Level"),
    ("summary", "Summary"),
    ("detection_count", "Detections"),
    ("reviewed", "Reviewed"),
]

# Extended columns for detailed exports
EXTENDED_EXPORT_COLUMNS: list[tuple[str, str]] = [
    *EXPORT_COLUMNS,
    ("object_types", "Object Types"),
    ("reasoning", "Reasoning"),
]


def format_export_value(row: EventExportRow, field: str) -> str:
    """Format a field value from an EventExportRow for export.

    Args:
        row: The export row
        field: Field name to format

    Returns:
        Formatted string value, sanitized for export safety
    """
    value = getattr(row, field, None)

    if value is None:
        return ""

    if isinstance(value, datetime):
        return value.isoformat()

    if isinstance(value, bool):
        return "Yes" if value else "No"

    if isinstance(value, int | float):
        return str(value)

    # String values get sanitized to prevent CSV injection
    return sanitize_export_value(str(value))


def generate_export_filename(prefix: str, export_format: ExportFormat) -> str:
    """Generate a timestamped filename for export.

    Args:
        prefix: Filename prefix (e.g., "events_export")
        export_format: Export format determining extension

    Returns:
        Filename with timestamp and appropriate extension
    """
    timestamp = datetime.now(UTC).strftime("%Y%m%d_%H%M%S")
    extension = EXPORT_EXTENSIONS[export_format]
    return f"{prefix}_{timestamp}{extension}"


def parse_accept_header(accept_header: str | None) -> ExportFormat:
    """Parse Accept header to determine export format.

    Supports content negotiation via standard HTTP Accept headers.
    Falls back to CSV format if header is missing or unrecognized.

    Args:
        accept_header: HTTP Accept header value

    Returns:
        ExportFormat enum value

    Examples:
        >>> parse_accept_header("text/csv")
        ExportFormat.CSV
        >>> parse_accept_header("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        ExportFormat.EXCEL
        >>> parse_accept_header(None)
        ExportFormat.CSV
    """
    if not accept_header:
        return ExportFormat.CSV

    # Parse accept header - may contain quality values like "text/csv;q=0.9"
    # We'll check each type in order of appearance
    for accept_type in accept_header.split(","):
        # Strip whitespace and quality value
        mime_type = accept_type.split(";")[0].strip().lower()

        if mime_type in ACCEPT_HEADER_MAPPING:
            return ACCEPT_HEADER_MAPPING[mime_type]

        # Check for wildcards
        if mime_type in {"*/*", "text/*"}:
            return ExportFormat.CSV

    # Default to CSV if no match
    return ExportFormat.CSV


def events_to_csv(
    events: Sequence[EventExportRow],
    columns: list[tuple[str, str]] | None = None,
) -> str:
    """Convert events to CSV format.

    Args:
        events: Sequence of EventExportRow objects
        columns: Optional list of (field_name, display_name) tuples.
                 Defaults to EXPORT_COLUMNS if not provided.

    Returns:
        CSV string with headers and data rows
    """
    if columns is None:
        columns = EXPORT_COLUMNS

    output = io.StringIO()
    writer = csv.writer(output)

    # Write header row
    header_row = [col[1] for col in columns]
    writer.writerow(header_row)

    # Write data rows
    for event in events:
        row = [format_export_value(event, col[0]) for col in columns]
        writer.writerow(row)

    return output.getvalue()


def events_to_csv_streaming(
    events: Sequence[EventExportRow],
    columns: list[tuple[str, str]] | None = None,
) -> Iterator[str]:
    """Generate CSV content as an async iterator for streaming.

    This is memory-efficient for large exports as it yields
    rows one at a time instead of building the entire CSV in memory.

    Args:
        events: Sequence of EventExportRow objects
        columns: Optional list of (field_name, display_name) tuples.

    Yields:
        CSV row strings (including newlines)
    """
    if columns is None:
        columns = EXPORT_COLUMNS

    output = io.StringIO()
    writer = csv.writer(output)

    # Yield header row
    header_row = [col[1] for col in columns]
    writer.writerow(header_row)
    yield output.getvalue()
    output.seek(0)
    output.truncate()

    # Yield data rows
    for event in events:
        row = [format_export_value(event, col[0]) for col in columns]
        writer.writerow(row)
        yield output.getvalue()
        output.seek(0)
        output.truncate()


def events_to_excel(
    events: Sequence[EventExportRow],
    columns: list[tuple[str, str]] | None = None,
    sheet_name: str = "Events",
) -> bytes:
    """Convert events to Excel (XLSX) format.

    Creates a professionally formatted Excel workbook with:
    - Styled header row (bold, background color, borders)
    - Auto-sized columns based on content width
    - Alternating row colors for readability
    - Proper column alignment based on data type

    Args:
        events: Sequence of EventExportRow objects
        columns: Optional list of (field_name, display_name) tuples.
                 Defaults to EXPORT_COLUMNS if not provided.
        sheet_name: Name for the worksheet (default: "Events")

    Returns:
        Excel file content as bytes
    """
    if columns is None:
        columns = EXPORT_COLUMNS

    # Create workbook and select active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name  # type: ignore[union-attr]

    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    alt_row_fill = PatternFill(start_color="E9EDF5", end_color="E9EDF5", fill_type="solid")

    # Track column widths for auto-sizing
    column_widths: list[int] = [len(col[1]) for col in columns]

    # Write header row
    for col_idx, (_, display_name) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=display_name)  # type: ignore[union-attr]
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Write data rows
    for row_idx, event in enumerate(events, start=2):
        for col_idx, (field_name, _) in enumerate(columns, start=1):
            value = format_export_value(event, field_name)

            # For Excel, we can use the raw value (not sanitized)
            # since openpyxl handles cell type properly
            raw_value = getattr(event, field_name, None)

            if isinstance(raw_value, datetime):
                # Excel/openpyxl doesn't support timezone-aware datetimes
                # Convert to naive datetime by removing tzinfo
                cell_value: Any = raw_value.replace(tzinfo=None)
            elif isinstance(raw_value, bool):
                cell_value = "Yes" if raw_value else "No"
            elif isinstance(raw_value, int | float):
                cell_value = raw_value  # Excel handles numbers natively
            elif raw_value is None:
                cell_value = ""
            else:
                # Sanitize string values for Excel too
                cell_value = sanitize_export_value(str(raw_value))

            cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)  # type: ignore[union-attr]
            cell.border = thin_border

            # Alternating row colors
            if row_idx % 2 == 0:
                cell.fill = alt_row_fill

            # Update column width tracking
            cell_len = len(str(value)) if value else 0
            if cell_len > column_widths[col_idx - 1]:
                column_widths[col_idx - 1] = min(cell_len, 50)  # Cap at 50 chars

    # Auto-size columns
    for col_idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width + 2  # type: ignore[union-attr]

    # Freeze header row
    ws.freeze_panes = "A2"  # type: ignore[union-attr]

    # Write to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


class ExportService:
    """Service for exporting event data in multiple formats.

    This service provides a high-level interface for exporting events,
    handling format selection, content generation, and response preparation.

    Supports background export jobs with progress tracking (NEM-1989).
    """

    def __init__(self, db: AsyncSession | None = None) -> None:
        """Initialize the export service.

        Args:
            db: Optional database session for querying events.
        """
        self._db = db
        # Ensure export directory exists
        EXPORT_DIR.mkdir(parents=True, exist_ok=True)

    def get_export_format(self, accept_header: str | None) -> ExportFormat:
        """Determine export format from Accept header.

        Args:
            accept_header: HTTP Accept header value

        Returns:
            ExportFormat enum value
        """
        return parse_accept_header(accept_header)

    def get_content_type(self, export_format: ExportFormat) -> str:
        """Get MIME type for export format.

        Args:
            export_format: Export format

        Returns:
            MIME type string
        """
        return EXPORT_MIME_TYPES[export_format]

    def get_filename(self, prefix: str, export_format: ExportFormat) -> str:
        """Generate export filename.

        Args:
            prefix: Filename prefix
            export_format: Export format

        Returns:
            Filename with timestamp and extension
        """
        return generate_export_filename(prefix, export_format)

    def export_events(
        self,
        events: Sequence[EventExportRow],
        export_format: ExportFormat,
        columns: list[tuple[str, str]] | None = None,
    ) -> bytes | str:
        """Export events in the specified format.

        Args:
            events: Sequence of EventExportRow objects
            export_format: Desired export format
            columns: Optional column definitions

        Returns:
            Export content as bytes (Excel) or string (CSV)
        """
        if export_format == ExportFormat.EXCEL:
            return events_to_excel(events, columns)
        else:
            return events_to_csv(events, columns)

    def get_content_disposition_header(
        self,
        filename: str,
        inline: bool = False,
    ) -> str:
        """Generate Content-Disposition header value.

        Args:
            filename: Export filename
            inline: If True, use inline disposition; otherwise attachment

        Returns:
            Content-Disposition header value
        """
        disposition = "inline" if inline else "attachment"
        return f'{disposition}; filename="{filename}"'

    async def export_events_with_progress(  # noqa: PLR0912
        self,
        job_id: str,
        job_tracker: JobTracker,
        export_format: str,
        camera_id: str | None = None,
        risk_level: str | None = None,
        start_date: str | None = None,
        end_date: str | None = None,
        reviewed: bool | None = None,
    ) -> dict[str, Any]:
        """Export events with progress tracking for background jobs.

        This method queries events from the database, exports them to a file,
        and updates the job tracker with progress information.

        Args:
            job_id: The job ID for progress tracking.
            job_tracker: JobTracker instance for progress updates.
            export_format: Export format ('csv', 'json', 'zip').
            camera_id: Optional camera ID filter.
            risk_level: Optional risk level filter.
            start_date: Optional start date filter (ISO format).
            end_date: Optional end date filter (ISO format).
            reviewed: Optional reviewed status filter.

        Returns:
            Dict with file_path, file_size, event_count, and format.

        Raises:
            ValueError: If database session not available.
        """
        from backend.models.camera import Camera
        from backend.models.event import Event

        if self._db is None:
            raise ValueError("Database session required for export_events_with_progress")

        # Build query
        query = select(Event).where(Event.deleted_at.is_(None))

        if camera_id is not None:
            query = query.where(Event.camera_id == camera_id)

        if risk_level is not None:
            query = query.where(Event.risk_level == risk_level)

        if start_date is not None:
            from datetime import datetime as dt

            start_dt = dt.fromisoformat(start_date.replace("Z", "+00:00"))
            query = query.where(Event.started_at >= start_dt)

        if end_date is not None:
            from datetime import datetime as dt

            end_dt = dt.fromisoformat(end_date.replace("Z", "+00:00"))
            query = query.where(Event.started_at <= end_dt)

        if reviewed is not None:
            query = query.where(Event.reviewed == reviewed)

        # Get total count first
        count_query = select(func.count()).select_from(query.subquery())
        result = await self._db.execute(count_query)
        total_count = result.scalar() or 0

        if total_count == 0:
            job_tracker.update_progress(job_id, 50, message="No events to export")
            # Create empty export file
            return await self._create_empty_export(export_format)

        job_tracker.update_progress(job_id, 10, message=f"Found {total_count} events to export")

        # Fetch events
        query = query.order_by(Event.started_at.desc())
        events_result = await self._db.execute(query)
        events: list[Event] = list(events_result.scalars().all())

        # Convert to export rows with progress
        export_rows: list[EventExportRow] = []
        for idx, event in enumerate(events):
            # Get camera name
            camera_name = "Unknown"
            if event.camera_id:
                camera_result = await self._db.execute(
                    select(Camera.name).where(Camera.id == event.camera_id)
                )
                camera_name = camera_result.scalar() or "Unknown"

            export_rows.append(
                EventExportRow(
                    event_id=event.id,
                    camera_name=camera_name,
                    started_at=event.started_at,
                    ended_at=event.ended_at,
                    risk_score=event.risk_score,
                    risk_level=event.risk_level,
                    summary=event.summary,
                    detection_count=event.detection_count or 0,
                    reviewed=event.reviewed or False,
                    object_types=event.object_types,
                    reasoning=event.reasoning,
                )
            )

            # Update progress every 100 events
            if (idx + 1) % PROGRESS_UPDATE_INTERVAL == 0:
                progress = 10 + int((idx + 1) / total_count * 70)  # 10-80% for fetching
                job_tracker.update_progress(
                    job_id, progress, message=f"Processing event {idx + 1}/{total_count}"
                )

        job_tracker.update_progress(job_id, 80, message=f"Writing {export_format.upper()} file...")

        # Generate export file
        timestamp = datetime.now(UTC).strftime("%Y%m%d_%H%M%S")

        if export_format == "csv":
            content = events_to_csv(export_rows, EXTENDED_EXPORT_COLUMNS)
            filename = f"events_export_{timestamp}.csv"
            filepath = EXPORT_DIR / filename
            filepath.write_text(content, encoding="utf-8")
        elif export_format == "json":
            content_dict = [
                {
                    "event_id": row.event_id,
                    "camera_name": row.camera_name,
                    "started_at": row.started_at.isoformat() if row.started_at else None,
                    "ended_at": row.ended_at.isoformat() if row.ended_at else None,
                    "risk_score": row.risk_score,
                    "risk_level": row.risk_level,
                    "summary": row.summary,
                    "detection_count": row.detection_count,
                    "reviewed": row.reviewed,
                    "object_types": row.object_types,
                    "reasoning": row.reasoning,
                }
                for row in export_rows
            ]
            filename = f"events_export_{timestamp}.json"
            filepath = EXPORT_DIR / filename
            filepath.write_text(json.dumps(content_dict, indent=2), encoding="utf-8")
        elif export_format == "zip":
            # Create JSON content and zip it
            content_dict = [
                {
                    "event_id": row.event_id,
                    "camera_name": row.camera_name,
                    "started_at": row.started_at.isoformat() if row.started_at else None,
                    "ended_at": row.ended_at.isoformat() if row.ended_at else None,
                    "risk_score": row.risk_score,
                    "risk_level": row.risk_level,
                    "summary": row.summary,
                    "detection_count": row.detection_count,
                    "reviewed": row.reviewed,
                    "object_types": row.object_types,
                    "reasoning": row.reasoning,
                }
                for row in export_rows
            ]
            filename = f"events_export_{timestamp}.zip"
            filepath = EXPORT_DIR / filename
            json_filename = f"events_export_{timestamp}.json"

            with zipfile.ZipFile(filepath, "w", zipfile.ZIP_DEFLATED) as zf:
                zf.writestr(json_filename, json.dumps(content_dict, indent=2))
        else:
            raise ValueError(f"Unsupported export format: {export_format}")

        job_tracker.update_progress(job_id, 95, message="Finalizing export...")

        file_size = filepath.stat().st_size

        logger.info(
            "Export completed",
            extra={
                "job_id": job_id,
                "format": export_format,
                "event_count": len(export_rows),
                "file_size": file_size,
            },
        )

        return {
            "file_path": f"/api/exports/{filename}",
            "file_size": file_size,
            "event_count": len(export_rows),
            "format": export_format,
        }

    async def _create_empty_export(
        self,
        export_format: str,
    ) -> dict[str, Any]:
        """Create an empty export file.

        Args:
            export_format: Export format.

        Returns:
            Export result dict.
        """
        timestamp = datetime.now(UTC).strftime("%Y%m%d_%H%M%S")

        if export_format == "csv":
            filename = f"events_export_{timestamp}.csv"
            filepath = EXPORT_DIR / filename
            # Write header only
            header = ",".join([col[1] for col in EXTENDED_EXPORT_COLUMNS])
            filepath.write_text(header + "\n", encoding="utf-8")
        elif export_format == "json":
            filename = f"events_export_{timestamp}.json"
            filepath = EXPORT_DIR / filename
            filepath.write_text("[]", encoding="utf-8")
        elif export_format == "zip":
            filename = f"events_export_{timestamp}.zip"
            filepath = EXPORT_DIR / filename
            json_filename = f"events_export_{timestamp}.json"
            with zipfile.ZipFile(filepath, "w", zipfile.ZIP_DEFLATED) as zf:
                zf.writestr(json_filename, "[]")
        else:
            raise ValueError(f"Unsupported export format: {export_format}")

        return {
            "file_path": f"/api/exports/{filename}",
            "file_size": filepath.stat().st_size,
            "event_count": 0,
            "format": export_format,
        }

    async def export_events_with_websocket(  # noqa: PLR0912
        self,
        progress_reporter: JobProgressReporter,
        export_format: str,
        camera_id: str | None = None,
        risk_level: str | None = None,
        start_date: str | None = None,
        end_date: str | None = None,
        reviewed: bool | None = None,
    ) -> dict[str, Any]:
        """Export events with WebSocket event emission for real-time progress.

        This method queries events from the database, exports them to a file,
        and emits WebSocket events for job progress. Uses JobProgressReporter
        to emit job.started, job.progress, job.completed, and job.failed events.

        The progress_reporter should NOT be started before calling this method -
        it will be started internally.

        Args:
            progress_reporter: JobProgressReporter instance for WebSocket events.
            export_format: Export format ('csv', 'json', 'zip').
            camera_id: Optional camera ID filter.
            risk_level: Optional risk level filter.
            start_date: Optional start date filter (ISO format).
            end_date: Optional end date filter (ISO format).
            reviewed: Optional reviewed status filter.

        Returns:
            Dict with file_path, file_size, event_count, format, and duration_seconds.

        Raises:
            ValueError: If database session not available.
        """
        from backend.models.camera import Camera
        from backend.models.event import Event

        if self._db is None:
            raise ValueError("Database session required for export_events_with_websocket")

        try:
            # Start the job and emit job.started event
            await progress_reporter.start(
                metadata={
                    "export_format": export_format,
                    "filters": {
                        "camera_id": camera_id,
                        "risk_level": risk_level,
                        "start_date": start_date,
                        "end_date": end_date,
                        "reviewed": reviewed,
                    },
                }
            )

            # Build query
            query = select(Event).where(Event.deleted_at.is_(None))

            if camera_id is not None:
                query = query.where(Event.camera_id == camera_id)

            if risk_level is not None:
                query = query.where(Event.risk_level == risk_level)

            if start_date is not None:
                from datetime import datetime as dt

                start_dt = dt.fromisoformat(start_date.replace("Z", "+00:00"))
                query = query.where(Event.started_at >= start_dt)

            if end_date is not None:
                from datetime import datetime as dt

                end_dt = dt.fromisoformat(end_date.replace("Z", "+00:00"))
                query = query.where(Event.started_at <= end_dt)

            if reviewed is not None:
                query = query.where(Event.reviewed == reviewed)

            # Get total count first
            count_query = select(func.count()).select_from(query.subquery())
            result = await self._db.execute(count_query)
            total_count = result.scalar() or 0

            await progress_reporter.report_progress(
                1, current_step=f"Found {total_count} events to export", force=True
            )

            if total_count == 0:
                # Create empty export file
                export_result = await self._create_empty_export(export_format)
                await progress_reporter.complete(
                    result_summary={
                        **export_result,
                        "message": "No events to export",
                    }
                )
                return export_result

            # Fetch events
            query = query.order_by(Event.started_at.desc())
            events_result = await self._db.execute(query)
            events: list[Event] = list(events_result.scalars().all())

            # Convert to export rows with progress
            export_rows: list[EventExportRow] = []
            for idx, event in enumerate(events):
                # Get camera name
                camera_name = "Unknown"
                if event.camera_id:
                    camera_result = await self._db.execute(
                        select(Camera.name).where(Camera.id == event.camera_id)
                    )
                    camera_name = camera_result.scalar() or "Unknown"

                export_rows.append(
                    EventExportRow(
                        event_id=event.id,
                        camera_name=camera_name,
                        started_at=event.started_at,
                        ended_at=event.ended_at,
                        risk_score=event.risk_score,
                        risk_level=event.risk_level,
                        summary=event.summary,
                        detection_count=event.detection_count or 0,
                        reviewed=event.reviewed or False,
                        object_types=event.object_types,
                        reasoning=event.reasoning,
                    )
                )

                # Report progress (throttled by reporter)
                # Use 10-80% range for fetching events
                progress_items = int((idx + 1) / total_count * 70)  # 0-70 range
                await progress_reporter.report_progress(
                    progress_items,
                    current_step=f"Processing event {idx + 1}/{total_count}",
                )

            # Writing file (80-95% range)
            await progress_reporter.report_progress(
                80, current_step=f"Writing {export_format.upper()} file...", force=True
            )

            # Generate export file
            timestamp = datetime.now(UTC).strftime("%Y%m%d_%H%M%S")

            if export_format == "csv":
                content = events_to_csv(export_rows, EXTENDED_EXPORT_COLUMNS)
                filename = f"events_export_{timestamp}.csv"
                filepath = EXPORT_DIR / filename
                filepath.write_text(content, encoding="utf-8")
            elif export_format == "json":
                content_dict = [
                    {
                        "event_id": row.event_id,
                        "camera_name": row.camera_name,
                        "started_at": row.started_at.isoformat() if row.started_at else None,
                        "ended_at": row.ended_at.isoformat() if row.ended_at else None,
                        "risk_score": row.risk_score,
                        "risk_level": row.risk_level,
                        "summary": row.summary,
                        "detection_count": row.detection_count,
                        "reviewed": row.reviewed,
                        "object_types": row.object_types,
                        "reasoning": row.reasoning,
                    }
                    for row in export_rows
                ]
                filename = f"events_export_{timestamp}.json"
                filepath = EXPORT_DIR / filename
                filepath.write_text(json.dumps(content_dict, indent=2), encoding="utf-8")
            elif export_format == "zip":
                content_dict = [
                    {
                        "event_id": row.event_id,
                        "camera_name": row.camera_name,
                        "started_at": row.started_at.isoformat() if row.started_at else None,
                        "ended_at": row.ended_at.isoformat() if row.ended_at else None,
                        "risk_score": row.risk_score,
                        "risk_level": row.risk_level,
                        "summary": row.summary,
                        "detection_count": row.detection_count,
                        "reviewed": row.reviewed,
                        "object_types": row.object_types,
                        "reasoning": row.reasoning,
                    }
                    for row in export_rows
                ]
                filename = f"events_export_{timestamp}.zip"
                filepath = EXPORT_DIR / filename
                json_filename = f"events_export_{timestamp}.json"

                with zipfile.ZipFile(filepath, "w", zipfile.ZIP_DEFLATED) as zf:
                    zf.writestr(json_filename, json.dumps(content_dict, indent=2))
            else:
                raise ValueError(f"Unsupported export format: {export_format}")

            await progress_reporter.report_progress(
                95, current_step="Finalizing export...", force=True
            )

            file_size = filepath.stat().st_size
            duration = progress_reporter.duration_seconds

            export_result = {
                "file_path": f"/api/exports/{filename}",
                "file_size": file_size,
                "event_count": len(export_rows),
                "format": export_format,
                "duration_seconds": duration,
            }

            logger.info(
                "Export completed with WebSocket events",
                extra={
                    "job_id": progress_reporter.job_id,
                    "format": export_format,
                    "event_count": len(export_rows),
                    "file_size": file_size,
                    "duration_seconds": duration,
                },
            )

            # Complete the job and emit job.completed event
            await progress_reporter.complete(result_summary=export_result)

            return export_result

        except Exception as e:
            # Fail the job and emit job.failed event
            await progress_reporter.fail(e, retryable=False)
            raise


# Global service instance
_export_service: ExportService | None = None


def get_export_service() -> ExportService:
    """Get or create the global ExportService instance.

    Returns:
        ExportService singleton instance
    """
    global _export_service  # noqa: PLW0603
    if _export_service is None:
        _export_service = ExportService()
    return _export_service


def reset_export_service() -> None:
    """Reset the global ExportService instance (for testing)."""
    global _export_service  # noqa: PLW0603
    _export_service = None
